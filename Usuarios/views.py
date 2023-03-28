import re
from django.shortcuts import get_object_or_404, render,redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User, Permission
from django.views import View
from Usuarios.models import PerfilUsuario
from django.contrib import messages
from validate_email_address import validate_email
from openpyxl import Workbook
from django.http import HttpResponse, JsonResponse
import pytz
from django.contrib.auth.decorators import permission_required
from django.core.paginator import Paginator
from django.db.models import Q
from .forms import *
from django.urls import reverse_lazy
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.utils.decorators import method_decorator


#Vistas basadas clases
class RegistroUsuarioView(CreateView):
    form_class = RegistroUsuarioForm
    template_name = 'Usuarios/registroform.html'
    success_url = reverse_lazy('lista')

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        tipo_usuario = self.request.user.perfilusuario.tipo_usuario
        if tipo_usuario == 'A' or self.request.user.is_superuser:
            form.fields.pop('creditos')
        return form

    def form_valid(self, form):
        # Obtener el tipo de usuario del usuario que está realizando el registro
        if self.request.user.is_superuser:
            tipo_usuario = 'A'
        elif self.request.user.perfilusuario.tipo_usuario == 'A':
            tipo_usuario = 'D'
        elif self.request.user.perfilusuario.tipo_usuario == 'D':
            tipo_usuario = 'C'
        #Verificar el correo
        email = form.cleaned_data.get('email')
        if User.objects.filter(email=email).exists():
            form.add_error('email', 'Este correo ya está registrado')
            return self.form_invalid(form)
        # Guardar el formulario
        user = form.save(self.request,tipo_usuario)

        # Asignar permisos al usuario registrado
        if tipo_usuario == 'D' or tipo_usuario == 'A':
            can_view_users_list_permission = Permission.objects.get(codename='can_view_users_list')
            user.user_permissions.add(can_view_users_list_permission)

        # Redireccionar al éxito
        messages.error(self.request, 'Usuario registrado', extra_tags='alert-succes')
        return redirect('lista')

@method_decorator(permission_required('Usuarios.can_view_users_list', raise_exception=True), name='dispatch')
class ListarUsuariosView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = User
    template_name = 'Usuarios/listado.html'
    context_object_name = 'usuarios'
    ordering = ['-date_joined']
    paginate_by = 5
    permission_required = 'Usuarios.can_view_users_list'
    login_url = reverse_lazy('home')

    def get_queryset(self):
        queryset = super().get_queryset().select_related('perfilusuario')
        user = self.request.user
        if user.perfilusuario.tipo_usuario == 'A' :
            queryset = User.objects.select_related('perfilusuario').all()
        elif user.perfilusuario.tipo_usuario == 'D' :
            queryset = User.objects.select_related('perfilusuario').filter(perfilusuario__registrado_por_id=user.id)
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        perm = self.request.user.has_perm('Usuarios.cambiar_credito')
        context['perm'] = perm
        return context

    def get(self, request, *args, **kwargs):
        if 'exportar' in request.GET:
            return self.exportar(request)
        elif 'suspender' in request.GET:
            return self.suspender(request)
        self.object_list = self.get_queryset()
        queryset = self.object_list
        nombre = request.GET.get('nombre')
        correo = request.GET.get('correo')
        fecha_desde = request.GET.get('fecha_desde')
        fecha_hasta = request.GET.get('fecha_hasta')

        if nombre:
            queryset = queryset.filter(Q(first_name__icontains=nombre))
        if correo:
            queryset = queryset.filter(Q(email__icontains=correo))
        if fecha_desde and fecha_hasta:
            queryset = queryset.filter(Q(date_joined__gte=fecha_desde) & Q(date_joined__lte=fecha_hasta))

        paginator = Paginator(queryset, self.paginate_by)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        context = self.get_context_data(page_obj=page_obj)
        return self.render_to_response(context)

    def exportar(request):
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="users.xlsx"'

        workbook = Workbook()

        # Selecciona la hoja activa
        worksheet = workbook.active
        worksheet.title = 'Usuarios'

        # Añade los encabezados
        columns = [
            'ID',
            'Nombre',
            'Apellido',
            'Correo electrónico',
            'Tipo de usuario',
            'Credito',
            'Activo',
            'Fecha de registro'
        ]
        row_num=1
        for col_num, column_title in enumerate(columns,1):
            cell = worksheet.cell(row=row_num,column=col_num)
            cell.value = column_title
        users = User.objects.select_related('perfilusuario').all()
        for user in users:
            row_num += 1
            tipo_usuario=user.perfilusuario.get_tipo_usuario_display()
            date_joined = user.date_joined.astimezone(pytz.timezone('UTC')).replace(tzinfo=None)
            if user.is_active:
                activo='si'
            else:
                activo='no'
            row = [
                user.id,
                user.first_name,
                user.last_name,
                user.email,
                tipo_usuario,
                user.perfilusuario.creditos,
                activo,
                date_joined
            ]
            for col_num, cell_value in enumerate(row,1):
                cell = worksheet.cell(row=row_num,column=col_num)
                cell.value = cell_value
        
        workbook.save(response)
        return response
    


@method_decorator(login_required(login_url='/login/'), name='dispatch')
@method_decorator(permission_required('Usuarios.can_edit_user', login_url='/'), name='dispatch')
class EditarUsuarioView(View):
    def get(self, request, id_usuario):
        usuario = get_object_or_404(User.objects.select_related('perfilusuario'), id=id_usuario)
        form = EditarUsuarioForm(instance=usuario)
        return render(request, 'Usuarios/editar.html', {'form': form, 'usuario': usuario})

    def post(self, request, id_usuario):
        usuario = get_object_or_404(User.objects.select_related('perfilusuario'), id=id_usuario)
        form = EditarUsuarioForm(request.POST, instance=usuario)
        if form.is_valid():
            form.save()
            messages.success(request, 'El usuario ha sido actualizado exitosamente.')
            return redirect('lista')
        else:
            messages.error(request, 'Por favor corrija los errores del formulario.', extra_tags='alert-danger')
            return render(request, 'Usuarios/editar.html', {'form': form, 'usuario': usuario})
        
@method_decorator(login_required(login_url='/login/'), name='dispatch')
class PerfilUpdateView(UpdateView):
    model = User
    form_class = PerfilForm
    template_name = 'Usuarios/perfilForm.html'
    success_url = reverse_lazy('perfil')

    def get_object(self, queryset=None):
        return self.request.user

@method_decorator(login_required(login_url='/login/'), name='dispatch')
@method_decorator(permission_required('Usuarios.can_view_users_list', login_url='/'), name='dispatch')
class EliminarUsuarioView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    model = User
    template_name = 'Usuarios/eliminar.html'
    success_url = reverse_lazy('lista')

    def get_object(self, queryset=None):
        usuario = get_object_or_404(User, pk=self.kwargs['id_usuario'])
        if usuario.perfilusuario.registrado_por_id == self.request.user.id:
            return usuario
        else:
            return None

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        usuario = self.get_object()
        if not usuario:
            context['error_message'] = 'No tienes permiso para eliminar este usuario.'
        return context



# Vistas con funciones.  
# @login_required(login_url='/login/')
# @permission_required('Usuarios.can_view_users_list', login_url='/')   
# def listar(request):
#     user=request.user 
#     if user.perfilusuario.tipo_usuario == 'A' :
#         usuarios = User.objects.select_related('perfilusuario').all()
#     elif user.perfilusuario.tipo_usuario == 'D' :
#         usuarios = User.objects.select_related('perfilusuario').filter(perfilusuario__registrado_por_id=user.id)
    
#     nombre = request.GET.get('nombre')
#     correo = request.GET.get('correo')
#     fecha_desde = request.GET.get('fecha_desde')
#     fecha_hasta = request.GET.get('fecha_hasta')

#     # Filtrar los usuarios en base a los parámetros de búsqueda
#     if nombre:
#         usuarios = usuarios.filter(Q(first_name__icontains=nombre))
    
#     if correo:
#         usuarios = usuarios.filter(Q(email__icontains=correo))
    
#     if fecha_desde and fecha_hasta:
#         usuarios = usuarios.filter(Q(date_joined__gte=fecha_desde) & Q(date_joined__lte=fecha_hasta))


#     paginator = Paginator(usuarios,5)
#     page_number=request.GET.get('page')
#     page_obj = paginator.get_page(page_number)
#     perm = request.user.has_perm('Usuarios.cambiar_credito')
#     context = {
#         'page_obj': page_obj,
#         'perm': perm,
#     }
#     return render(request,'Usuarios/listado.html',context)
       

# @login_required(login_url='/login/')
# @permission_required('Usuarios.can_view_users_list',login_url='/')   
# def editar(request,id_usuario):
#     usuario = User.objects.select_related('perfilusuario').get(id=id_usuario)
#     if usuario.perfilusuario.registrado_por_id == request.user.id:
#         if request.method == 'POST':
#             usuario.first_name=request.POST['nombre']
#             usuario.last_name=request.POST['apellidos']
#             usuario.email=request.POST['correo']
#             usuario.username=request.POST['username']
#             usuario.save()
#         return render(request,'Usuarios/editar.html',{'usuario':usuario})
#     else:
#         messages.error(request, 'No tienes permiso para editar este usuario.', extra_tags='alert-danger')
#         return redirect('lista')

# @login_required(login_url='/login/')
# @permission_required('Usuarios.can_view_users_list',login_url='/')
# def eliminar(request,id_usuario):
#     usuario = get_object_or_404(User, pk=id_usuario)
#     if usuario.perfilusuario.registrado_por_id == request.user.id:
#         usuario.delete()
#     else:
#         return render(request, 'Usuarios/listado.html', {'error_message': 'no tienes permiso para eliminar este usuario.'})
#     return redirect('lista')

# @login_required(login_url='/login/')
# @permission_required('Usuarios.can_view_users_list',login_url='/')
# def registro_form(request):
#     if request.method == 'POST':
#         form = RegistroUsuarioForm(request.POST)
#         if form.is_valid():
#             if request.user.is_superuser:
#                 tipo_usuario = 'A'
#             elif request.user.perfilusuario.tipo_usuario == 'A' :
#                 tipo_usuario = 'D'
#             elif request.user.perfilusuario.tipo_usuario == 'D' :
#                 tipo_usuario='C'
#             form.save(request,tipo_usuario)
#             return redirect('lista')
#     else:
#         tipo_usuario=request.user.perfilusuario.tipo_usuario
#         form = RegistroUsuarioForm()
#         if tipo_usuario == 'A' or request.user.is_superuser:
#             form.fields.pop('creditos')
#     return render(request, 'Usuarios/registroform.html', {'form': form})

# def registro(request):
#     user=request.user
#     if not user.perfilusuario.tipo_usuario == 'C':
#         if request.method == 'POST':
#             registrando=request.user
#             registrando=registrando.id
#             correo = request.POST['correo']
#             username=request.POST['username']
#             password = request.POST['password']
#             nombre = request.POST['nombre']
#             apellidos = request.POST['apellidos']
#             credito = request.POST['credito']
#             if not validate_email(correo):
#                 return render(request, 'Usuarios/registro.html', {'error_message': 'El correo electrónico es invalido'})
#             if user.is_superuser:
#                 tipo_usuario == 'A'
#             elif user.perfilusuario.tipo_usuario == 'A' :
#                 tipo_usuario = 'D'
#             elif user.perfilusuario.tipo_usuario == 'D' :
#                 tipo_usuario='C'
#             if User.objects.filter(email=correo).exists():    
#                 return render(request, 'Usuarios/registro.html', {'error_message': 'El correo electrónico ya están en uso'})
#             if not re.search("^(?=.*[0-9])(?=.*[!@#$%^&*-_.])[a-zA-Z0-9!@#$%^&*]{10,}$", password):
#                 return render(request, 'Usuarios/registro.html', {'error_message': 'La contraseña debe tener al menos 10 caracteres y contener números y caracteres especiales(!@#$%^&*-_.)'})
#             user = User.objects.create_user(username=username,password=password,first_name=nombre,last_name=apellidos,email=correo)
#             perfil= PerfilUsuario(user=user)
#             perfil.tipo_usuario=tipo_usuario
#             perfil.creditos=credito
#             perfil.registra=registrando
#             perfil.save()
#             if user.perfilusuario.tipo_usuario == 'D' or user.perfilusuario.tipo_usuario == 'A':
#                 can_view_users_list_permission = Permission.objects.get(codename='can_view_users_list')
#                 user.user_permissions.add(can_view_users_list_permission)
#             return redirect('registrar')
#         return render(request,'Usuarios/registro.html')
#     else:
#         messages.warning(request, "No tienes permiso para registrar nuevos usuarios")
#         return redirect('home')

# @login_required(login_url='/login/')
# def user_detail(request):
#     form = perfilForm(instance=request.user)
#     return render(request, 'Usuarios/perfilForm.html', {'form': form})

# @login_required(login_url='/login/')
# def perfil(request):
#     usuario=request.user
#     if request.method == 'POST':
#         usuario.first_name=request.POST['nombre']
#         usuario.last_name=request.POST['apellidos']
#         usuario.email=request.POST['correo']
#         usuario.save()
#     return render(request,'Usuarios/perfilCliente.html')



# @login_required(login_url='/login/')
# @permission_required('Usuarios.can_view_users_list',login_url='/')
# def exportar(request):
#     response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#     response['Content-Disposition'] = 'attachment; filename="users.xlsx"'

#     workbook = Workbook()

#     # Selecciona la hoja activa
#     worksheet = workbook.active
#     worksheet.title = 'Usuarios'

#     # Añade los encabezados
#     columns = [
#         'ID',
#         'Nombre',
#         'Apellido',
#         'Correo electrónico',
#         'Tipo de usuario',
#         'Credito',
#         'Activo',
#         'Fecha de registro'
#     ]
#     row_num=1
#     for col_num, column_title in enumerate(columns,1):
#         cell = worksheet.cell(row=row_num,column=col_num)
#         cell.value = column_title
#     users = User.objects.select_related('perfilusuario').all()
#     for user in users:
#         row_num += 1
#         tipo_usuario=user.perfilusuario.get_tipo_usuario_display()
#         date_joined = user.date_joined.astimezone(pytz.timezone('UTC')).replace(tzinfo=None)
#         if user.is_active:
#             activo='si'
#         else:
#             activo='no'
#         row = [
#             user.id,
#             user.first_name,
#             user.last_name,
#             user.email,
#             tipo_usuario,
#             user.perfilusuario.creditos,
#             activo,
#             date_joined
#         ]
#         for col_num, cell_value in enumerate(row,1):
#             cell = worksheet.cell(row=row_num,column=col_num)
#             cell.value = cell_value
    
#     workbook.save(response)
#     return response

def get_credits(request):
    user = request.user
    return JsonResponse({'credits': user.perfilusuario.creditos})

@login_required(login_url='/login/')
@permission_required('Usuarios.cambiar_credito', login_url='/')       
def editar_credito(request, cliente_id):
    cliente = get_object_or_404(User, id=cliente_id)
    if request.method == 'POST':
        cliente.perfilusuario.creditos = request.POST['credito']
        cliente.save()
        return render(request, 'Usuarios/editar_credito.html', {'cliente': cliente})
    else:
        return render(request, 'Usuarios/editar_credito.html', {'cliente': cliente})
    
def suspender(request):
    usuario_id = request.GET.get('usuario_id')
    if request.GET.get('activo') == 'true':
        activo=1
    else:
        activo=0
    print(usuario_id)
    print(activo)
    usuario = User.objects.get(id=usuario_id)
    print(usuario.first_name)
    usuario.is_active = activo
    usuario.save()
    return JsonResponse({'success': True})